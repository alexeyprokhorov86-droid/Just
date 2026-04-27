#!/usr/bin/env python3
"""Применяет ручную верификацию из identity_review.xlsx к km_entities.

Читает столбцы M (OK/NO) и N (комментарий/правильный employee_ref_key) и:
  OK     → km_entities.confidence = 1.0, attrs.verified_at = now()
  NO + комментарий с ФИО → перевешивает TG на правильную km_entity
  NO + «внешний», «не нужен» → удаляет TG из текущей km_entity
  NO + «не в 1С / не заведён» → оставляет, но помечает attrs.not_in_1c=true

После применения — пересчитывает downstream (comm_users.km_entity_id /
employee_ref_key, email_employee_mapping.tg_user_id) от km_entities.

Запуск:
  python3 apply_identity_review.py /home/admin/telegram_logger_bot/.tmp_input/identity_review.xlsx
  python3 apply_identity_review.py path/to/file --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import pathlib
import re
import sys
from datetime import datetime

import psycopg2
import requests
from openpyxl import load_workbook
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv

REPO = pathlib.Path(__file__).resolve().parent
load_dotenv(REPO / ".env")

ROUTER_AI_URL = os.getenv("ROUTERAI_BASE_URL", "https://routerai.ru/api/v1")
ROUTER_AI_KEY = os.getenv("ROUTERAI_API_KEY")
LLM_MODEL = "anthropic/claude-opus-4.7"

DB_HOST = os.getenv("DB_HOST", "172.20.0.2")
DB_NAME = os.getenv("DB_NAME", "knowledge_base")
DB_USER = os.getenv("DB_USER", "knowledge")
DB_PASS = os.getenv("DB_PASSWORD")


def log(m: str): print(f"[{datetime.now().strftime('%H:%M:%S')}] {m}", flush=True)


def conn_db():
    return psycopg2.connect(host=DB_HOST, dbname=DB_NAME, user=DB_USER, password=DB_PASS)


# ─────────────────────────────────────────────────────────────────────
#  LLM helper — find employee by free-form name
# ─────────────────────────────────────────────────────────────────────

def llm_find_employee(name_text: str, employees: list[dict]) -> str | None:
    """name_text — комментарий пользователя ('в 1с как Прохорова Ирина').
    Возвращает employee.ref_key или None."""
    if not ROUTER_AI_KEY:
        return None
    cands = "\n".join(f"  - {e['ref_key']}: {e['name']}" for e in employees)
    prompt = f"""Из текста выдели имя человека и сматчи с одним из сотрудников 1С.

Текст пользователя: {name_text!r}

Кандидаты (ref_key: ФИО):
{cands}

Учитывай транслит и инициалы. Если уверен — верни один ref_key.
Если не уверен или нет в списке — верни NONE.
Ответ — только ref_key или NONE, без пояснений."""
    try:
        r = requests.post(
            f"{ROUTER_AI_URL}/chat/completions",
            headers={"Authorization": f"Bearer {ROUTER_AI_KEY}"},
            json={"model": LLM_MODEL,
                  "messages": [{"role": "user", "content": prompt}],
                  "temperature": 0, "max_tokens": 64},
            timeout=60,
        )
        ans = r.json()["choices"][0]["message"]["content"].strip()
        if ans == "NONE": return None
        if re.match(r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", ans):
            return ans
        return None
    except Exception as e:
        log(f"  LLM err: {e}"); return None


# ─────────────────────────────────────────────────────────────────────
#  Apply review
# ─────────────────────────────────────────────────────────────────────

EXTERNAL_KEYWORDS = ("внешн", "не в 1с", "не нужен", "не заведен", "не заведён",
                    "бывш", "нет в группах", "неизвестно", "не играет роли")


def classify_comment(comment: str) -> str:
    c = (comment or "").lower()
    if any(k in c for k in ("внешн", "не нужен", "неизвестно", "не играет роли")):
        return "external"
    if "бывш" in c or "нет в группах" in c:
        return "former"
    if "не в 1с" in c or "не заведен" in c or "не заведён" in c:
        return "not_in_1c"
    if re.search(r"в 1[сc] как ", c) or re.search(r"\b[А-ЯЁ][а-яё]+ [А-ЯЁ][а-яё]", comment or ""):
        return "rematch"
    if c.strip():
        return "other_comment"
    return "no_comment"


def apply_review(xlsx_path: str, dry: bool) -> dict:
    log(f"reading {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["TG → 1С"]

    conn = conn_db()
    # Active employees as candidates for LLM
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute("SELECT ref_key, name FROM c1_employees "
                    "WHERE COALESCE(is_archived,false)=false AND name IS NOT NULL")
        employees = [dict(r) for r in cur.fetchall()]
    log(f"  employees-candidates: {len(employees)}")

    stats = {"ok": 0, "rematch": 0, "external": 0, "former": 0, "not_in_1c": 0,
             "skipped": 0, "errors": 0}

    for r in range(2, ws.max_row + 1):
        tg_id = ws.cell(row=r, column=1).value
        display_name = ws.cell(row=r, column=2).value
        cur_km = ws.cell(row=r, column=7).value
        cur_emp_name = ws.cell(row=r, column=8).value
        ok = (ws.cell(row=r, column=13).value or "").strip()
        comment = (ws.cell(row=r, column=14).value or "").strip()
        if not tg_id:
            continue

        if ok.upper() == "OK":
            if cur_km and not dry:
                with conn.cursor() as cur:
                    cur.execute("""
                        UPDATE km_entities SET
                          confidence=1.0,
                          attrs = jsonb_set(
                                    jsonb_set(COALESCE(attrs,'{}'::jsonb),
                                              '{verified_at}', to_jsonb(NOW()::text)),
                                    '{verified_by}', to_jsonb('admin'::text)),
                          updated_at=NOW()
                        WHERE id=%s
                    """, (cur_km,))
                conn.commit()
            stats["ok"] += 1
            continue

        kind = classify_comment(comment)

        if kind == "external":
            # Удалить TG из текущей km_entity (она не должна на этого сотрудника указывать)
            log(f"  tg={tg_id} ({display_name}) → EXTERNAL: {comment[:60]}")
            if not dry and cur_km:
                with conn.cursor() as cur:
                    cur.execute("""
                        UPDATE km_entities SET
                          attrs = jsonb_set(
                            COALESCE(attrs,'{}'::jsonb), '{tg_user_ids}',
                            ((SELECT COALESCE(jsonb_agg(v), '[]'::jsonb)
                              FROM jsonb_array_elements(COALESCE(attrs->'tg_user_ids','[]'::jsonb)) v
                              WHERE v <> to_jsonb(%s::bigint)))
                          ),
                          updated_at = NOW()
                        WHERE id = %s
                    """, (tg_id, cur_km))
                    cur.execute(
                        "UPDATE comm_users SET is_external=true, km_entity_id=NULL, "
                        "  employee_ref_key=NULL WHERE tg_user_id=%s", (tg_id,))
                conn.commit()
            stats["external"] += 1

        elif kind == "former":
            # Сотрудник был, но больше не в группах. Не отвязываем от km_entity (история нужна),
            # просто пометим attrs.former_in_tg=true.
            log(f"  tg={tg_id} ({display_name}) → FORMER: {comment[:60]}")
            if not dry and cur_km:
                with conn.cursor() as cur:
                    cur.execute("""
                        UPDATE km_entities SET
                          attrs = jsonb_set(COALESCE(attrs,'{}'::jsonb),
                                  '{former_in_tg}', 'true'::jsonb),
                          updated_at = NOW()
                        WHERE id = %s
                    """, (cur_km,))
                conn.commit()
            stats["former"] += 1

        elif kind == "not_in_1c":
            # Сотрудник реальный, но в 1С его нет — пометим km_entity (если есть) как not_in_1c
            log(f"  tg={tg_id} ({display_name}) → NOT_IN_1C: {comment[:60]}")
            if not dry and cur_km:
                with conn.cursor() as cur:
                    cur.execute("""
                        UPDATE km_entities SET
                          attrs = jsonb_set(COALESCE(attrs,'{}'::jsonb),
                                  '{not_in_1c}', 'true'::jsonb),
                          updated_at = NOW()
                        WHERE id = %s
                    """, (cur_km,))
                conn.commit()
            stats["not_in_1c"] += 1

        elif kind == "rematch":
            # Перевесить TG на правильного сотрудника по комментарию
            new_ref = llm_find_employee(comment, employees)
            if not new_ref:
                log(f"  tg={tg_id} ({display_name}) → REMATCH ! LLM не нашёл: {comment[:60]}")
                stats["errors"] += 1
                continue
            log(f"  tg={tg_id} ({display_name}) → REMATCH к {new_ref}")
            if not dry:
                with conn.cursor() as cur:
                    # Удалить TG из старой km_entity
                    if cur_km:
                        cur.execute("""
                            UPDATE km_entities SET
                              attrs = jsonb_set(
                                COALESCE(attrs,'{}'::jsonb), '{tg_user_ids}',
                                ((SELECT COALESCE(jsonb_agg(v), '[]'::jsonb)
                                  FROM jsonb_array_elements(COALESCE(attrs->'tg_user_ids','[]'::jsonb)) v
                                  WHERE v <> to_jsonb(%s::bigint)))
                              ),
                              updated_at = NOW()
                            WHERE id = %s
                        """, (tg_id, cur_km))
                    # Найти/обновить новый km_entity
                    cur.execute("SELECT id FROM km_entities WHERE entity_type='person' "
                                "AND source_ref=%s", (new_ref,))
                    new_km = cur.fetchone()
                    if new_km:
                        new_km_id = new_km[0]
                        cur.execute("""
                            UPDATE km_entities SET
                              confidence=1.0, updated_at=NOW(),
                              attrs = jsonb_set(
                                jsonb_set(
                                  jsonb_set(
                                    COALESCE(attrs,'{}'::jsonb),
                                    '{tg_user_ids}',
                                    COALESCE(attrs->'tg_user_ids','[]'::jsonb) || to_jsonb(%s::bigint)),
                                  '{verified_at}', to_jsonb(NOW()::text)),
                                '{verified_by}', to_jsonb('admin'::text))
                            WHERE id = %s
                        """, (tg_id, new_km_id))
                        cur.execute(
                            "UPDATE comm_users SET km_entity_id=%s, employee_ref_key=%s "
                            "WHERE tg_user_id=%s",
                            (new_km_id, new_ref, tg_id))
                conn.commit()
            stats["rematch"] += 1

        else:
            stats["skipped"] += 1

    # Dedup tg_user_ids (на случай если задвоилось)
    if not dry:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE km_entities SET attrs = jsonb_set(
                  attrs, '{tg_user_ids}',
                  (SELECT COALESCE(jsonb_agg(DISTINCT v), '[]'::jsonb)
                   FROM jsonb_array_elements(attrs->'tg_user_ids') v))
                WHERE entity_type='person'
                  AND attrs ? 'tg_user_ids'
                  AND jsonb_array_length(attrs->'tg_user_ids') > 1
            """)
        conn.commit()

    conn.close()
    return stats


def main():
    p = argparse.ArgumentParser()
    p.add_argument("xlsx_path")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()
    if not pathlib.Path(args.xlsx_path).exists():
        print(f"File not found: {args.xlsx_path}", file=sys.stderr); return 1
    stats = apply_review(args.xlsx_path, args.dry_run)
    log(f"stats: {stats}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
